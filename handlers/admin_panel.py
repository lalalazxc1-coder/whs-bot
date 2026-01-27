from aiogram import Router, F, types
from aiogram.fsm.context import FSMContext
from aiogram.filters import Command
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.types import FSInputFile
import openpyxl
import os
from datetime import datetime

import config
import database.requests as db
from states import AdminPanelState, AdminBranchState, AdminItemState, AdminContactState

router = Router()

def get_admin_main_kb(problems_count: int = 0, questions_count: int = 0, orders_count: int = 0, branches_count: int = 0, items_count: int = 0, contacts_count: int = 0):
    builder = InlineKeyboardBuilder()
    
    builder.button(text="📋 Управление отчетами", callback_data="admin_manage_reports")
    builder.button(text=f"📦 Заказы ({orders_count})", callback_data="admin_show_orders")
    builder.button(text=f"⚠️ Проблемы ({problems_count})", callback_data="admin_show_tickets_problem")
    builder.button(text=f"❓ Вопросы ({questions_count})", callback_data="admin_show_tickets_question")
    
    # Management
    builder.button(text=f"🏢 Филиалы ({branches_count})", callback_data="admin_branches_menu")
    builder.button(text=f"📦 Товары ({items_count})", callback_data="admin_items_menu")
    builder.button(text=f"📞 Контакты ({contacts_count})", callback_data="admin_contacts")
    
    builder.button(text="📢 Рассылка объявлений", callback_data="admin_broadcast")
    builder.adjust(1)
    return builder.as_markup()

def get_admin_reports_management_kb(inventory_open: bool):
    builder = InlineKeyboardBuilder()
    
    # Toggle button
    if inventory_open:
        builder.button(text="🛑 Остановить сбор", callback_data="admin_inventory_toggle")
    else:
        builder.button(text="▶️ Начать сбор", callback_data="admin_inventory_toggle")
        
    builder.button(text="📊 Прогресс сдачи", callback_data="admin_reports_progress")
    builder.button(text="⚙️ Авто-расписание", callback_data="admin_auto_schedule") # New
    builder.button(text="⬇️ Отчеты Excel", callback_data="admin_reports_menu")
    
    if inventory_open:
        builder.button(text="📢 Напомнить должникам", callback_data="admin_remind_debtors")
        
    builder.button(text="⬅️ Назад", callback_data="admin_cancel")
    builder.adjust(1)
    return builder.as_markup()

@router.callback_query(F.data == "admin_manage_reports")
async def admin_manage_reports_handler(callback: types.CallbackQuery):
    if callback.from_user.id not in config.ADMIN_IDS: return
    
    inventory_open = await db.is_inventory_open()
    status_text = "🟢 Сбор отчетов ОТКРЫТ" if inventory_open else "🔴 Сбор отчетов ЗАКРЫТ"
    
    await callback.message.edit_text(
        f"📋 **Управление отчетами**\n\nСтатус: {status_text}",
        reply_markup=get_admin_reports_management_kb(inventory_open)
    )

@router.callback_query(F.data == "admin_show_orders")
async def admin_show_orders_list(callback: types.CallbackQuery):
    if callback.from_user.id not in config.ADMIN_IDS: return

    # Заказы = problems + startswith
    all_problems = await db.get_open_tickets(ticket_type="problem")
    orders = [t for t in all_problems if t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    
    if not orders:
        await callback.answer("Нет открытых заказов.", show_alert=True)
        return
        
    text = f"📦 **Список заказов:**\n\n"
    for t in orders:
        clean_msg = t.message.replace("[ЗАКАЗ МАТЕРИАЛОВ]", "").strip()
        text += f"🆔 `#{t.id}` | {t.created_at.strftime('%d.%m %H:%M')}\n"
        text += f"👤 {t.user_name} ({t.branch_name})\n"
        text += f"🛒 {clean_msg}\n"
        text += f"-------------------------\n"
        
    if len(text) > 4000:
        text = text[:4000] + "\n...(обрезано)..."
        
    builder = InlineKeyboardBuilder()
    builder.button(text="✍️ Ответить на тикет", callback_data="admin_reply_ticket_start")
    builder.button(text="⬅️ Назад", callback_data="admin_cancel")
    builder.adjust(1)
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup(), parse_mode="Markdown")

# --- Auto Schedule Settings ---

@router.callback_query(F.data == "admin_auto_schedule")
async def admin_auto_schedule_menu(callback: types.CallbackQuery):
    if callback.from_user.id not in config.ADMIN_IDS: return
    
    # Load settings
    auto_mode = await db.get_setting("inventory_auto_mode", "0") == "1"
    start_day = await db.get_setting("inventory_start_day", "25")
    end_day = await db.get_setting("inventory_end_day", "1")
    
    mode_icon = "✅ Включено" if auto_mode else "🔴 Выключено"
    
    text = (
        f"⚙️ **Авто-расписание инвентаризации**\n\n"
        f"Статус: **{mode_icon}**\n"
        f"📅 День начала (открытие): **{start_day}-е число**\n"
        f"📅 День окончания (закрытие): **{end_day}-е число**\n\n"
        f"_В указанный день начала бот автоматически откроет сбор и разошлет уведомления._"
    )
    
    builder = InlineKeyboardBuilder()
    toggle_text = "🔴 Выключить" if auto_mode else "🟢 Включить"
    builder.button(text=toggle_text, callback_data="admin_auto_toggle")
    builder.button(text="✏️ Изм. день начала", callback_data="admin_auto_set_start")
    builder.button(text="✏️ Изм. день конца", callback_data="admin_auto_set_end")
    builder.button(text="⬅️ Назад", callback_data="admin_manage_reports")
    builder.adjust(1)
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup(), parse_mode="Markdown")

@router.callback_query(F.data == "admin_auto_toggle")
async def admin_auto_toggle(callback: types.CallbackQuery):
    current = await db.get_setting("inventory_auto_mode", "0")
    new_val = "0" if current == "1" else "1"
    await db.set_setting("inventory_auto_mode", new_val)
    await admin_auto_schedule_menu(callback)

@router.callback_query(F.data == "admin_auto_set_start")
async def admin_auto_set_start(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введите день месяца (число от 1 до 31), когда нужно **ОТКРЫВАТЬ** сбор:")
    await state.set_state(AdminPanelState.auto_start_day)

@router.message(AdminPanelState.auto_start_day)
async def admin_auto_save_start(message: types.Message, state: FSMContext):
    if not message.text.isdigit() or not (1 <= int(message.text) <= 31):
        await message.answer("❌ Введите число от 1 до 31.")
        return
        
    await db.set_setting("inventory_start_day", message.text)
    
    # Fetch counts for menu
    all_problems = await db.get_open_tickets(ticket_type="problem")
    orders = [t for t in all_problems if t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    real_problems = [t for t in all_problems if not t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    p_count = len(real_problems)
    o_count = len(orders)
    q_count = await db.count_tickets(ticket_type="question", status="open")
    b_count = await db.count_branches()
    i_count = await db.count_active_items()
    c_count = await db.count_contacts()

    await message.answer(f"✅ День начала установлен на {message.text}-е число.", reply_markup=get_admin_main_kb(p_count, q_count, o_count, b_count, i_count, c_count))
    await state.clear()
    
    # Simple redirect back to menu via text message isn't great, let's show menu
    # But we can't edit last message easily from here without storing ID. 
    # Just show 'Saved' and maybe the menu again.
    
    # Let's clean up user message? No permissions usually. 
    # Just resend the menu logic?
    # Simulating callback for menu is hard from message.
    # Let's just send text.

@router.callback_query(F.data == "admin_auto_set_end")
async def admin_auto_set_end(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введите день месяца (число от 1 до 31), когда нужно **ЗАКРЫВАТЬ** сбор:")
    await state.set_state(AdminPanelState.auto_end_day)

@router.message(AdminPanelState.auto_end_day)
async def admin_auto_save_end(message: types.Message, state: FSMContext):
    if not message.text.isdigit() or not (1 <= int(message.text) <= 31):
        await message.answer("❌ Введите число от 1 до 31.")
        return
        
    await db.set_setting("inventory_end_day", message.text)
    
    # Fetch counts
    all_problems = await db.get_open_tickets(ticket_type="problem")
    orders = [t for t in all_problems if t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    real_problems = [t for t in all_problems if not t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    p_count = len(real_problems)
    o_count = len(orders)
    q_count = await db.count_tickets(ticket_type="question", status="open")
    b_count = await db.count_branches()
    i_count = await db.count_active_items()
    c_count = await db.count_contacts()
    
    await message.answer(f"✅ День окончания установлен на {message.text}-е число.", reply_markup=get_admin_main_kb(p_count, q_count, o_count, b_count, i_count, c_count))
    await state.clear()

@router.callback_query(F.data == "admin_inventory_toggle")
async def admin_inventory_toggle_handler(callback: types.CallbackQuery):
    if callback.from_user.id not in config.ADMIN_IDS: return
    
    current_status = await db.is_inventory_open()
    new_status = not current_status
    
    await db.set_setting("inventory_open", "1" if new_status else "0")
    
    if new_status:
        # Авто-уведомление
        await callback.message.edit_text("⏳ Сбор открыт! Рассылаю уведомления...")
        users = await db.get_all_users()
        count = 0
        for u in users:
            try:
                # Текст в зависимости от языка
                lang = u.language if u.language else "ru"
                
                if lang == "kz":
                    msg = "🔔 **Назар аударыңыз!**\n\nҚалдықтарды жинау басталды. Есеп тапсырыңыз."
                else:
                     msg = "🔔 **Внимание!**\n\nОткрыт сбор отчетов по остаткам. Пожалуйста, сдайте отчет."
                     
                await callback.message.bot.send_message(u.telegram_id, msg)
                count += 1
            except: pass
        
        await callback.answer(f"✅ Уведомление отправлено {count} сотр.")
    else:
        await callback.answer("🔴 Сбор закрыт!")

    # Refresh menu
    await admin_manage_reports_handler(callback)

# Progress Handler
@router.callback_query(F.data == "admin_reports_progress")
async def admin_reports_progress_handler(callback: types.CallbackQuery):
    users = await db.get_all_users()
    # Get reports for last 24h
    reports = await db.get_reports_by_range(1)
    
    submitted_ids = {r.user_id for r in reports}
    
    # Group by branch
    branches = {}
    for u in users:
        b_name = u.branch.name if u.branch else "Без филиала"
        if b_name not in branches:
            branches[b_name] = {"submitted": [], "pending": []}
            
        # У нас пока нет имени в таблице User, используем ID
        user_label = f"ID {u.telegram_id}"
        
        if u.telegram_id in submitted_ids:
            branches[b_name]["submitted"].append(user_label)
        else:
            branches[b_name]["pending"].append(user_label)
            
    text = "📊 **Прогресс сдачи (24ч):**\n\n"
    for b_name, data in branches.items():
        sub = len(data['submitted'])
        pen = len(data['pending'])
        total = sub + pen
        text += f"🏢 **{b_name}**: {sub}/{total}\n"
        if pen > 0:
             text += f"⚠️ Не сдали: {pen} чел.\n"
        text += "\n"
        
    await callback.message.edit_text(text, reply_markup=InlineKeyboardBuilder().button(text="⬅️ Назад", callback_data="admin_manage_reports").as_markup(), parse_mode="Markdown")

@router.callback_query(F.data == "admin_remind_debtors")
async def admin_remind_debtors_handler(callback: types.CallbackQuery):
    await callback.message.edit_text("⏳ Рассылаю уведомления должникам...")
    users = await db.get_users_pending_report()
    count = 0
    for u in users:
        try:
             # Локализуем и тут
             lang = u.language if u.language else "ru"
             if lang == "kz":
                 msg = "🔔 Ескерту: Есеп тапсырыңыз!"
             else:
                 msg = "🔔 Напоминание: Пожалуйста, сдайте отчет!"
                 
             await callback.message.bot.send_message(u.telegram_id, msg)
             count += 1
        except: pass
    
    # Возвращаем меню
    inventory_open = await db.is_inventory_open()
    status_text = "🟢 Сбор отчетов ОТКРЫТ" if inventory_open else "🔴 Сбор отчетов ЗАКРЫТ"
    
    await callback.message.edit_text(
        f"✅ Напоминание отправлено {count} пользователям.\n\n📋 **Управление отчетами**\nСтатус: {status_text}",
        reply_markup=get_admin_reports_management_kb(inventory_open)
    )

@router.callback_query(F.data == "notify_inventory_start")
async def notify_inventory_start(callback: types.CallbackQuery):
    users = await db.get_all_users()
    count = 0
    await callback.message.edit_text("⏳ Рассылаю уведомления...")
    
    for u in users:
        try:
            # Текст в зависимости от языка
            lang = u.language if u.language else "ru"
            # Хардкод для скорости
            if lang == "kz":
                msg = "🔔 **Назар аударыңыз!**\n\nҚалдықтарды жинау басталды. «📦 Қалдықтарды жіберу» батырмасын басыңыз."
            else:
                msg = "🔔 **Внимание!**\n\nОткрыт сбор отчетов по остаткам. Пожалуйста, нажмите кнопку «📦 Отправить остатки»."
                
            await callback.message.bot.send_message(u.telegram_id, msg)
            count += 1
        except: pass
        
    await callback.message.answer(f"✅ Уведомление отправлено {count} пользователям.")

def get_admin_reports_kb():
    builder = InlineKeyboardBuilder()
    builder.button(text="📅 За 7 дней", callback_data="admin_export_7")
    builder.button(text="📅 За 30 дней", callback_data="admin_export_30")
    builder.button(text="♾ За все время", callback_data="admin_export_0")
    builder.button(text="⬅️ Назад", callback_data="admin_cancel")
    builder.adjust(1)
    return builder.as_markup()

@router.message(Command("admin"))
async def cmd_admin_panel(message: types.Message):
    if message.from_user.id not in config.ADMIN_IDS:
        return

    # Загрузим статистику
    users_count = await db.count_users()
    reports_today = await db.count_reports(days=1)
    
    # Logic to separate orders from problems
    all_problems = await db.get_open_tickets(ticket_type="problem")
    orders = [t for t in all_problems if t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    real_problems = [t for t in all_problems if not t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    
    p_count = len(real_problems)
    o_count = len(orders)
    q_count = await db.count_tickets(ticket_type="question", status="open")
    
    # New counts
    b_count = await db.count_branches()
    i_count = await db.count_active_items()
    c_count = await db.count_contacts()
    
    text = (
        f"🛠 **Панель Администратора**\n\n"
        f"👥 Пользователей: `{users_count}`\n"
        f"📋 Отчетов за 24ч: `{reports_today}`"
    )

    await message.answer(text, reply_markup=get_admin_main_kb(p_count, q_count, o_count, b_count, i_count, c_count), parse_mode="Markdown")

@router.callback_query(F.data == "admin_cancel")
async def admin_cancel(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    
    # Reload stats to show fresh main menu
    users_count = await db.count_users()
    reports_today = await db.count_reports(days=1)
    
    # Logic to separate orders from problems
    all_problems = await db.get_open_tickets(ticket_type="problem")
    orders = [t for t in all_problems if t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    real_problems = [t for t in all_problems if not t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    
    p_count = len(real_problems)
    o_count = len(orders)
    q_count = await db.count_tickets(ticket_type="question", status="open")
    
    # New counts
    b_count = await db.count_branches()
    i_count = await db.count_active_items()
    c_count = await db.count_contacts()
    
    text = (
        f"🛠 **Панель Администратора**\n\n"
        f"👥 Пользователей: `{users_count}`\n"
        f"📋 Отчетов за 24ч: `{reports_today}`"
    )
    
    await callback.message.edit_text(text, reply_markup=get_admin_main_kb(p_count, q_count, o_count, b_count, i_count, c_count))

@router.callback_query(F.data.startswith("admin_show_tickets_"))
async def admin_show_tickets_list(callback: types.CallbackQuery):
    if callback.from_user.id not in config.ADMIN_IDS: return

    t_type = callback.data.split("_")[3] # problem or question
    
    tickets = await db.get_open_tickets(ticket_type=t_type)
    
    # If type is problem, exclude orders!
    if t_type == "problem":
        tickets = [t for t in tickets if not t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    
    if not tickets:
        await callback.answer(f"Нет открытых тикетов типа '{t_type}'.", show_alert=True)
        return
        
    text = f"📂 **Список тикетов ({t_type}):**\n\n"
    for t in tickets:
        text += f"🆔 `#{t.id}` | {t.created_at.strftime('%d.%m %H:%M')}\n"
        text += f"👤 {t.user_name} ({t.branch_name})\n"
        text += f"💬 {t.message[:100]}...\n"
        text += f"-------------------------\n"
        
    # Split text if too long (simple check)
    if len(text) > 4000:
        text = text[:4000] + "\n...(обрезано)..."
        
    builder = InlineKeyboardBuilder()
    builder.button(text="✍️ Ответить на тикет", callback_data="admin_reply_ticket_start")
    builder.button(text="⬅️ Назад", callback_data="admin_cancel")
    builder.adjust(1)
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup(), parse_mode="Markdown")

@router.callback_query(F.data == "admin_reply_ticket_start")
async def admin_reply_ticket_start(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer("✍️ Введите ID тикета (число после #):")
    await state.set_state(AdminPanelState.ticket_reply_id)

@router.message(AdminPanelState.ticket_reply_id)
async def admin_reply_id_input(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("❌ ID должен быть числом. Попробуйте снова:")
        return
        
    tid = int(message.text)
    ticket = await db.get_ticket(tid)
    
    if not ticket:
        await message.answer("❌ Тикет с таким ID не найден.")
        return
        
    if ticket.status != "open":
         await message.answer("⚠️ Этот тикет уже закрыт.")
         return
         
    await state.update_data(reply_tid=tid)
    await message.answer(f"💬 Введите ответ для пользователя {ticket.user_name}:")
    await state.set_state(AdminPanelState.ticket_reply_msg)

@router.message(AdminPanelState.ticket_reply_msg)
async def admin_reply_msg_input(message: types.Message, state: FSMContext):
    data = await state.get_data()
    tid = data.get("reply_tid")
    reply_text = message.text
    
    # Сохраняем и закрываем
    await db.close_ticket(
        ticket_id=tid, 
        reply_text=reply_text, 
        responder_id=message.from_user.id, 
        responder_name=message.from_user.full_name
    )
    
    # Уведомляем пользователя
    ticket = await db.get_ticket(tid) # reload to be sure
    try:
        user_lang = "ru" # TODO: fetch user language if possible, or store in ticket
        # пока на русском
        await message.bot.send_message(
            ticket.user_id,
            f"📩 **Ответ на ваш тикет #{tid}:**\n\n{reply_text}"
        )
    except:
        await message.answer("⚠️ Не удалось отправить уведомление пользователю (бот заблокирован?). ответа сохранен в БД.")
        
    # Обновляем счетчики для меню
    all_problems = await db.get_open_tickets(ticket_type="problem")
    orders = [t for t in all_problems if t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    real_problems = [t for t in all_problems if not t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    
    p_count = len(real_problems)
    o_count = len(orders)
    q_count = await db.count_tickets(ticket_type="question", status="open")
    
    b_count = await db.count_branches()
    i_count = await db.count_active_items()
    c_count = await db.count_contacts()
    
    await message.answer(f"✅ Ответ отправлен! Тикет #{tid} закрыт.", reply_markup=get_admin_main_kb(p_count, q_count, o_count, b_count, i_count, c_count))
    await state.clear()

# --- Рассылка ---

@router.callback_query(F.data == "admin_broadcast")
async def broadcast_start(callback: types.CallbackQuery, state: FSMContext):
    if callback.from_user.id not in config.ADMIN_IDS: return
    
    # Выбор цели: Все или Филиал
    builder = InlineKeyboardBuilder()
    builder.button(text="🌍 Всем", callback_data="broadcast_target_all")
    
    branches = await db.get_branches()
    for b in branches:
        builder.button(text=f"🏢 {b.name}", callback_data=f"broadcast_target_branch_{b.id}")
        
    builder.button(text="❌ Отмена", callback_data="admin_cancel")
    builder.adjust(2)
    
    await callback.message.edit_text("📢 **Рассылка:** Кому отправить сообщение?", reply_markup=builder.as_markup())
    await state.set_state(AdminPanelState.select_target)

@router.callback_query(AdminPanelState.select_target, F.data.startswith("broadcast_target_"))
async def broadcast_enter_msg(callback: types.CallbackQuery, state: FSMContext):
    target = callback.data.split("_")[2] # all or branch
    branch_id = None
    if target == "branch":
        branch_id = int(callback.data.split("_")[3])
        
    await state.update_data(target=target, branch_id=branch_id)
    
    await callback.message.edit_text("✍️ Введите текст объявления (можно с фото):")
    await state.set_state(AdminPanelState.broadcast_msg)

@router.message(AdminPanelState.broadcast_msg)
async def broadcast_send(message: types.Message, state: FSMContext):
    data = await state.get_data()
    target = data.get("target")
    branch_id = data.get("branch_id")
    
    # Получаем пользователей
    if target == "all":
        users = await db.get_all_users()
    else:
        users = await db.get_users_by_branch(branch_id)
        
    if not users:
        await message.answer("Пользователи не найдены.")
        await state.clear()
        return
    
    count = 0
    # Отправка
    notify_msg = await message.answer(f"⏳ Начинаю рассылку для {len(users)} пользователей...")
    
    for u in users:
        try:
            await message.copy_to(u.telegram_id)
            count += 1
        except Exception:
            pass # Юзер заблочил бота
            
    await notify_msg.edit_text(f"✅ Рассылка завершена!\nДоставлено: {count} из {len(users)}")
    await state.clear()

@router.callback_query(F.data == "admin_cancel")
async def admin_cancel(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("Отменено.", reply_markup=get_admin_main_kb())

# --- Экспорт ---

@router.callback_query(F.data == "admin_reports_menu")
async def export_menu_handler(callback: types.CallbackQuery):
    await callback.message.edit_text("📊 **Выберите период отчета:**", reply_markup=get_admin_reports_kb())

@router.callback_query(F.data.startswith("admin_export_"))
async def export_data_handler(callback: types.CallbackQuery):
    if callback.from_user.id not in config.ADMIN_IDS: return
    
    days = int(callback.data.split("_")[2]) # 7, 30, 0
    period_name = f"{days} дней" if days > 0 else "Все время"
    
    await callback.message.answer(f"⏳ Генерирую отчет ({period_name})...")
    
    reports = await db.get_reports_by_range(days)
    all_tickets = await db.get_tickets_by_range(days)
    
    # Разделяем тикеты на Обычные и Заказы
    # Сначала убираем заказы
    non_orders = [t for t in all_tickets if not t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    orders = [t for t in all_tickets if t.message.startswith("[ЗАКАЗ МАТЕРИАЛОВ]")]
    
    # Теперь разделяем non_orders на Проблемы и Вопросы
    problems = [t for t in non_orders if t.ticket_type == 'problem']
    questions = [t for t in non_orders if t.ticket_type == 'question']
    # Для совместимости старых записей (где ticket_type может быть не заполнен, но дефолт 'problem')
    # можно считать problem по умолчанию. Но так как мы уже добавили колонку с дефолтом, все ок.
    
    wb = openpyxl.Workbook()
    
    # --- Лист 1: Инвентаризация ---
    ws1 = wb.active
    ws1.title = "Inventory"
    ws1.append(["ID", "Date", "Branch", "Sector", "User", "Report Data"])
    
    for r in reports:
        # User Link Formula
        user_display = r.user_name if r.user_name else str(r.user_id)
        user_link_formula = f'=HYPERLINK("tg://user?id={r.user_id}", "{user_display}")'
        
        # Sector (handle None for old records)
        sector_display = r.sector if r.sector else "N/A"
        
        ws1.append([r.id, r.timestamp, r.branch_name, sector_display, user_link_formula, r.report_data])

    # --- Лист 2: Проблемы (Problems) ---
    ws2 = wb.create_sheet("Problems")
    ws2.append(["ID", "Date", "Status", "Branch", "Sender", "Message", "Responder", "Reply", "Reply Date"])
    
    for t in problems:
        # Sender Link
        sender_display = t.user_name if t.user_name else str(t.user_id)
        sender_link = f'=HYPERLINK("tg://user?id={t.user_id}", "{sender_display}")'
        
        # Responder Link
        if t.responder_id:
            responder_display = t.responder_name if t.responder_name else str(t.responder_id)
            responder_link = f'=HYPERLINK("tg://user?id={t.responder_id}", "{responder_display}")'
        else:
            responder_link = ""

        ws2.append([
            t.id, t.created_at, t.status, t.branch_name, 
            sender_link, t.message, 
            responder_link, t.reply_message, t.reply_at
        ])

    # --- Лист 3: Вопросы (Questions) ---
    ws_q = wb.create_sheet("Questions")
    ws_q.append(["ID", "Date", "Status", "Branch", "Sender", "Message", "Responder", "Reply", "Reply Date"])
    
    for t in questions:
        # Sender Link
        sender_display = t.user_name if t.user_name else str(t.user_id)
        sender_link = f'=HYPERLINK("tg://user?id={t.user_id}", "{sender_display}")'
        
        # Responder Link
        if t.responder_id:
            responder_display = t.responder_name if t.responder_name else str(t.responder_id)
            responder_link = f'=HYPERLINK("tg://user?id={t.responder_id}", "{responder_display}")'
        else:
            responder_link = ""

        ws_q.append([
            t.id, t.created_at, t.status, t.branch_name, 
            sender_link, t.message, 
            responder_link, t.reply_message, t.reply_at
        ])
        
    # --- Лист 3: Заявки (Orders) ---
    ws3 = wb.create_sheet("Orders")
    ws3.append(["ID", "Date", "Status", "Branch", "User", "Order Details", "Responder", "Note"])
    
    for o in orders:
        # Убираем префикс [ЗАКАЗ МАТЕРИАЛОВ] для красоты
        clean_msg = o.message.replace("[ЗАКАЗ МАТЕРИАЛОВ]", "").strip()
        
        user_display = o.user_name if o.user_name else str(o.user_id)
        user_link = f'=HYPERLINK("tg://user?id={o.user_id}", "{user_display}")'
        
        responder_link = ""
        if o.responder_id:
             responder_display = o.responder_name if o.responder_name else str(o.responder_id)
             responder_link = f'=HYPERLINK("tg://user?id={o.responder_id}", "{responder_display}")'
             
        ws3.append([
            o.id, o.created_at, o.status, o.branch_name,
            user_link, clean_msg,
            responder_link, o.reply_message
        ])
        
    # Сохраняем
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    
    # Отправляем
    try:
        file = FSInputFile(filename)
        await callback.message.answer_document(file, caption=f"📊 Отчет ({period_name})")
    finally:
        if os.path.exists(filename):
            os.remove(filename)

    await callback.answer()

# --- Управление контактами (UI) ---

# --- Управление контактами (UI) ---

@router.callback_query(F.data == "admin_contacts")
async def admin_contacts_list(callback: types.CallbackQuery, state: FSMContext = None):
    # state for manual call support
    if callback.from_user.id not in config.ADMIN_IDS: return

    contacts = await db.get_contacts()
    
    builder = InlineKeyboardBuilder()
    
    # Список как кнопок
    for c in contacts:
        label = f"🏢 {c.department} | {c.info[:15]}"
        builder.button(text=label, callback_data=f"admin_contact_sel_{c.id}")
        
    builder.button(text="➕ Добавить контакт", callback_data="admin_add_contact")
    builder.button(text="⬅️ Назад", callback_data="admin_cancel")
    builder.adjust(1)
    
    msg_text = f"📞 **Управление контактами ({len(contacts)} шт.):**\nВыберите контакт для управления."
    
    if isinstance(callback, types.CallbackQuery):
        await callback.message.edit_text(msg_text, reply_markup=builder.as_markup(), parse_mode="Markdown")
    else:
        await callback.answer(msg_text, reply_markup=builder.as_markup(), parse_mode="Markdown")

@router.callback_query(F.data.startswith("admin_contact_sel_"))
async def admin_contact_select(callback: types.CallbackQuery):
    cid = int(callback.data.split("_")[3])
    contact = await db.get_contact(cid)
    
    if not contact:
        await callback.answer("Контакт не найден.", show_alert=True)
        await admin_contacts_list(callback)
        return
        
    builder = InlineKeyboardBuilder()
    builder.button(text="✏️ Редактировать", callback_data=f"admin_contact_edit_{cid}")
    builder.button(text="🗑 Удалить", callback_data=f"admin_contact_del_{cid}")
    builder.button(text="⬅️ Назад", callback_data="admin_contacts")
    builder.adjust(1)
    
    text = f"👤 **Контакт:**\n🏢 Отдел: `{contact.department}`\nℹ️ Данные: `{contact.info}`"
    await callback.message.edit_text(text, reply_markup=builder.as_markup(), parse_mode="Markdown")

# --- Добавление ---
@router.callback_query(F.data == "admin_add_contact")
async def admin_add_contact_start(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("✍️ Введите название отдела (например: Логистика):")
    await state.set_state(AdminContactState.add_dept)

@router.message(AdminContactState.add_dept)
async def admin_add_contact_dept(message: types.Message, state: FSMContext):
    await state.update_data(new_dept=message.text)
    await message.answer("✍️ Введите данные контакта (Имя, Телефон, Почта):")
    await state.set_state(AdminContactState.add_info)

@router.message(AdminContactState.add_info)
async def admin_add_contact_save(message: types.Message, state: FSMContext):
    data = await state.get_data()
    dept = data.get("new_dept")
    info = message.text
    
    await db.add_contact(dept, info)
    await message.answer(f"✅ Контакт добавлен: **{dept}** - {info}", parse_mode="Markdown")
    await admin_contacts_list(message)
    await state.clear()

# --- Редактирование ---
@router.callback_query(F.data.startswith("admin_contact_edit_"))
async def admin_contact_edit_start(callback: types.CallbackQuery, state: FSMContext):
    cid = int(callback.data.split("_")[3])
    contact = await db.get_contact(cid)
    
    if not contact:
        await callback.answer("Ошибка доступа.", show_alert=True)
        return
        
    await state.update_data(edit_cid=cid)
    await callback.message.edit_text(f"✍️ Сейчас отдел: `{contact.department}`.\nВведите новое название отдела (или отправьте точку '.' чтобы оставить текущее):", parse_mode="Markdown")
    await state.set_state(AdminContactState.edit_dept)

@router.message(AdminContactState.edit_dept)
async def admin_contact_edit_dept(message: types.Message, state: FSMContext):
    new_dept = message.text
    data = await state.get_data()
    cid = data.get("edit_cid")
    
    # Optional: fetch old if dot
    contact = await db.get_contact(cid)
    if new_dept == ".":
        new_dept = contact.department
        
    await state.update_data(edit_dept=new_dept)
    await message.answer(f"✍️ Сейчас данные: `{contact.info}`.\nВведите новые данные (или '.' чтобы оставить):", parse_mode="Markdown")
    await state.set_state(AdminContactState.edit_info)

@router.message(AdminContactState.edit_info)
async def admin_contact_edit_save(message: types.Message, state: FSMContext):
    new_info = message.text
    data = await state.get_data()
    cid = data.get("edit_cid")
    dept = data.get("edit_dept")
    
    contact = await db.get_contact(cid)
    if new_info == ".":
        new_info = contact.info
        
    await db.update_contact(cid, dept, new_info)
    await message.answer("✅ Контакт обновлен.")
    await admin_contacts_list(message)
    await state.clear()

# --- Удаление ---
@router.callback_query(F.data.startswith("admin_contact_del_"))
async def admin_del_contact(callback: types.CallbackQuery):
    cid = int(callback.data.split("_")[3])
    if await db.delete_contact(cid):
        await callback.answer("✅ Контакт удален")
        await admin_contacts_list(callback)
    else:
        await callback.answer("❌ Ошибка при удалении", show_alert=True)

# --- Управление Филиалами (Branches) ---

@router.callback_query(F.data == "admin_branches_menu")
async def admin_branches_menu(callback: types.CallbackQuery, state: FSMContext = None):
    # state argument is for manual calling
    branches = await db.get_branches()
    
    builder = InlineKeyboardBuilder()
    
    for b in branches:
        builder.button(text=f"🏢 {b.name}", callback_data=f"admin_branch_sel_{b.id}")
        
    builder.button(text="➕ Добавить филиал", callback_data="admin_branch_add")
    builder.button(text="⬅️ Назад", callback_data="admin_cancel")
    builder.adjust(1)
    
    msg_text = "🏢 **Управление филиалами:**\nВыберите филиал для редактирования или создайте новый."
    
    # Check if this is a message or callback (since we might call from save handler)
    if isinstance(callback, types.CallbackQuery):
        await callback.message.edit_text(msg_text, reply_markup=builder.as_markup(), parse_mode="Markdown")
    else:
        # It's a Message
        await callback.answer(msg_text, reply_markup=builder.as_markup(), parse_mode="Markdown")

@router.callback_query(F.data == "admin_branch_add")
async def admin_branch_add_start(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("✍️ Введите название нового филиала:")
    await state.set_state(AdminBranchState.add_name)

@router.message(AdminBranchState.add_name)
async def admin_branch_save_new(message: types.Message, state: FSMContext):
    name = message.text
    await db.add_branch(name)
    await message.answer(f"✅ Филиал **{name}** создан.")
    # Show menu again
    await admin_branches_menu(message) # Passing message as callback arg hack
    await state.clear()

@router.callback_query(F.data.startswith("admin_branch_sel_"))
async def admin_branch_select(callback: types.CallbackQuery):
    bid = int(callback.data.split("_")[3])
    branch = await db.get_branch_by_id(bid)
    
    if not branch:
        await callback.answer("Филиал не найден.", show_alert=True)
        await admin_branches_menu(callback)
        return
        
    builder = InlineKeyboardBuilder()
    builder.button(text="✏️ Переименовать", callback_data=f"admin_branch_edit_{bid}")
    builder.button(text="🗑 Удалить", callback_data=f"admin_branch_del_{bid}")
    builder.button(text="⬅️ Назад", callback_data="admin_branches_menu")
    builder.adjust(1)
    
    await callback.message.edit_text(f"🏢 Филиал: **{branch.name}**\nВыберите действие:", reply_markup=builder.as_markup(), parse_mode="Markdown")

@router.callback_query(F.data.startswith("admin_branch_edit_"))
async def admin_branch_edit_start(callback: types.CallbackQuery, state: FSMContext):
    bid = int(callback.data.split("_")[3])
    await state.update_data(editing_bid=bid)
    await callback.message.edit_text("✍️ Введите новое название:")
    await state.set_state(AdminBranchState.edit_name)

@router.message(AdminBranchState.edit_name)
async def admin_branch_save_edit(message: types.Message, state: FSMContext):
    data = await state.get_data()
    bid = data.get("editing_bid")
    new_name = message.text
    
    if await db.rename_branch(bid, new_name):
        await message.answer(f"✅ Филиал переименован в **{new_name}**.")
    else:
        await message.answer("❌ Ошибка при обновлении.")
        
    await admin_branches_menu(message)
    await state.clear()

@router.callback_query(F.data.startswith("admin_branch_del_"))
async def admin_branch_delete_handler(callback: types.CallbackQuery):
    bid = int(callback.data.split("_")[3])
    
    # Confirm? For speed let's just delete or use simple confirm.
    # User asked for "create, edit, delete", usually implies ability to do so.
    # Adding confirmation is safer.
    
    if await db.delete_branch(bid):
        await callback.answer("✅ Филиал удален.")
        await admin_branches_menu(callback)
    else:
        await callback.answer("❌ Ошибка. Возможно, есть привязанные пользователи.", show_alert=True)

# --- Управление Товарами (Items) ---

@router.callback_query(F.data == "admin_items_menu")
async def admin_items_menu(callback: types.CallbackQuery, state: FSMContext = None):
    items = await db.get_active_items()
    
    builder = InlineKeyboardBuilder()
    
    # Товаров может быть много, лучше список с прокруткой или просто длинный список.
    # Пока делаем полный список, aiogram сам разобьет если >100.
    # Лучше 2 в ряд
    for item in items:
        builder.button(text=f"{item.name}", callback_data=f"admin_item_sel_{item.id}")
        
    builder.button(text="➕ Добавить товар", callback_data="admin_item_add")
    builder.button(text="⬅️ Назад", callback_data="admin_cancel")
    builder.adjust(2)
    
    msg_text = f"📦 **Управление товарами ({len(items)} шт.):**\nВыберите товар для редактирования или создайте новый."
    
    if isinstance(callback, types.CallbackQuery):
        await callback.message.edit_text(msg_text, reply_markup=builder.as_markup(), parse_mode="Markdown")
    else:
        await callback.answer(msg_text, reply_markup=builder.as_markup(), parse_mode="Markdown")

@router.callback_query(F.data == "admin_item_add")
async def admin_item_add_start(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("✍️ Введите название нового товара:")
    await state.set_state(AdminItemState.add_name)

@router.message(AdminItemState.add_name)
async def admin_item_save_new(message: types.Message, state: FSMContext):
    name = message.text
    # Проверка на пустой ввод?
    if not name or len(name) < 2:
        await message.answer("Слишком короткое название.")
        return
        
    await db.add_item(name)
    await message.answer(f"✅ Товар **{name}** создан.")
    await admin_items_menu(message)
    await state.clear()

@router.callback_query(F.data.startswith("admin_item_sel_"))
async def admin_item_select(callback: types.CallbackQuery):
    item_id = int(callback.data.split("_")[3])
    item = await db.get_item(item_id)
    
    if not item:
        await callback.answer("Товар не найден (возможно удален).", show_alert=True)
        await admin_items_menu(callback)
        return
        
    builder = InlineKeyboardBuilder()
    builder.button(text="✏️ Переименовать", callback_data=f"admin_item_edit_{item_id}")
    builder.button(text="🗑 Удалить", callback_data=f"admin_item_del_{item_id}")
    builder.button(text="⬅️ Назад", callback_data="admin_items_menu")
    builder.adjust(1)
    
    await callback.message.edit_text(f"📦 Товар: **{item.name}**\nВыберите действие:", reply_markup=builder.as_markup(), parse_mode="Markdown")

@router.callback_query(F.data.startswith("admin_item_edit_"))
async def admin_item_edit_start(callback: types.CallbackQuery, state: FSMContext):
    item_id = int(callback.data.split("_")[3])
    await state.update_data(editing_item_id=item_id)
    await callback.message.edit_text("✍️ Введите новое название товара:")
    await state.set_state(AdminItemState.edit_name)

@router.message(AdminItemState.edit_name)
async def admin_item_save_edit(message: types.Message, state: FSMContext):
    data = await state.get_data()
    item_id = data.get("editing_item_id")
    new_name = message.text
    
    if await db.rename_item(item_id, new_name):
        await message.answer(f"✅ Товар переименован в **{new_name}**.")
    else:
        await message.answer("❌ Ошибка при обновлении.")
        
    await admin_items_menu(message)
    await state.clear()

@router.callback_query(F.data.startswith("admin_item_del_"))
async def admin_item_delete_handler(callback: types.CallbackQuery):
    item_id = int(callback.data.split("_")[3])
    
    if await db.delete_item(item_id):
        await callback.answer("✅ Товар удален (скрыт).")
        await admin_items_menu(callback)
    else:
        await callback.answer("❌ Ошибка при удалении.", show_alert=True)
